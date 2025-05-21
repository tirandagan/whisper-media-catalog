import os
import logging
import pandas as pd
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.orm import sessionmaker
import datetime
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

def export_database_to_excel(db_path, export_folder=None):
    """
    Export all tables in the SQLite database to Excel
    
    Args:
        db_path: Path to the SQLite database
        export_folder: Folder to save the Excel file (defaults to database folder)
        
    Returns:
        str: Path to the exported Excel file
    """
    try:
        # Create the export folder path if not provided
        if not export_folder:
            export_folder = os.path.dirname(db_path)
        
        # Only use a single filename without timestamp
        excel_filename = "video_database.xlsx"
        excel_path = os.path.join(export_folder, excel_filename)
        
        # Connect to the database
        engine = create_engine(f"sqlite:///{db_path}")
        inspector = inspect(engine)
        
        # Create a Excel writer object
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Create and populate metadata sheet first
            create_metadata_sheet(writer)
            
            # Export each table to a separate worksheet
            for table_name in inspector.get_table_names():
                # Skip SQLite internal tables and association tables
                if table_name.startswith('sqlite_') or table_name == 'transcription_keywords':
                    continue
                
                # Read the table into a pandas DataFrame
                # Use direct SQL query to ensure all columns (including recently added ones) are included
                query = f"SELECT * FROM {table_name}"
                df = pd.read_sql_query(query, engine)
                
                # Format date columns
                for col in df.columns:
                    if col.endswith('_at') and df[col].dtype == 'object':
                        df[col] = pd.to_datetime(df[col])
                
                # Write the DataFrame to Excel
                df.to_excel(writer, sheet_name=table_name, index=False)
                
                # Format the sheet
                format_sheet(writer.sheets[table_name], df, table_name)
                
                # Log which columns are present in the table only at debug level
                logger.debug(f"Exported table {table_name} with {len(df)} rows and columns: {', '.join(df.columns)}")
            
            # Create a simple joined view directly with pandas
            try:
                # First, create a view with video and transcription data
                base_query = """
                SELECT v.*, t.is_transcribed, t.transcribed_at, t.transcript_text, 
                       t.transcript_file, t.suggested_title, t.summary
                FROM videos v
                LEFT JOIN transcriptions t ON v.id = t.video_id
                ORDER BY 
                    CASE 
                        WHEN v.status = 'Missing' THEN 1
                        WHEN v.status = 'Error Transcribing' THEN 2
                        WHEN v.status = 'New' THEN 3
                        WHEN v.status = 'Transcribed' THEN 4
                        ELSE 5
                    END,
                    v.filename
                """
                joined_df = pd.read_sql_query(base_query, engine)
                
                # Add keywords as a comma-separated list using a subquery
                if 'transcription_keywords' in inspector.get_table_names() and 'keywords' in inspector.get_table_names():
                    # Create a separate DataFrame with transcription_id -> keywords mapping
                    try:
                        # First try with GROUP_CONCAT which is SQLite-specific
                        keywords_query = """
                        SELECT t.id AS transcription_id, GROUP_CONCAT(k.name, ', ') AS keywords
                        FROM transcriptions t
                        LEFT JOIN transcription_keywords tk ON t.id = tk.transcription_id
                        LEFT JOIN keywords k ON tk.keyword_id = k.id
                        GROUP BY t.id
                        """
                        keywords_df = pd.read_sql_query(keywords_query, engine)
                    except Exception as e:
                        logger.debug(f"GROUP_CONCAT not supported, using alternative approach: {str(e)}")
                        # Fallback if GROUP_CONCAT is not supported
                        # Get raw data and do the aggregation in pandas
                        base_keywords_query = """
                        SELECT t.id AS transcription_id, k.name AS keyword
                        FROM transcriptions t
                        LEFT JOIN transcription_keywords tk ON t.id = tk.transcription_id
                        LEFT JOIN keywords k ON tk.keyword_id = k.id
                        WHERE k.name IS NOT NULL
                        """
                        raw_keywords_df = pd.read_sql_query(base_keywords_query, engine)
                        
                        # Aggregate in pandas
                        if not raw_keywords_df.empty:
                            keywords_df = raw_keywords_df.groupby('transcription_id')['keyword'].apply(
                                lambda x: ', '.join(x)
                            ).reset_index(name='keywords')
                        else:
                            # Create empty dataframe with correct columns
                            keywords_df = pd.DataFrame(columns=['transcription_id', 'keywords'])
                    
                    # Add a keywords column to the joined dataframe using a left join
                    if not keywords_df.empty:
                        # Merge on transcription ID
                        # First, create the transcription_id column in the joined_df
                        video_ids = joined_df['id'].tolist()
                        transcription_ids_query = f"""
                        SELECT v.id AS video_id, t.id AS transcription_id
                        FROM videos v
                        LEFT JOIN transcriptions t ON v.id = t.video_id
                        WHERE v.id IN ({','.join(['?' for _ in video_ids])})
                        """
                        conn = sqlite3.connect(db_path)
                        id_mapping_df = pd.read_sql_query(transcription_ids_query, conn, params=video_ids)
                        conn.close()
                        
                        # Add transcription_id to joined_df
                        joined_df = pd.merge(
                            joined_df, 
                            id_mapping_df, 
                            left_on='id', 
                            right_on='video_id', 
                            how='left'
                        )
                        
                        # Now join with keywords
                        joined_df = pd.merge(
                            joined_df, 
                            keywords_df, 
                            left_on='transcription_id', 
                            right_on='transcription_id', 
                            how='left'
                        )
                        
                        # Drop the temporary columns - safely check if columns exist first
                        columns_to_drop = []
                        if 'transcription_id' in joined_df.columns:
                            columns_to_drop.append('transcription_id')
                        if 'video_id_y' in joined_df.columns:
                            columns_to_drop.append('video_id_y')
                            # Only rename if we had video_id_y (meaning we also have video_id_x)
                            joined_df = joined_df.rename(columns={'video_id_x': 'video_id'})
                            
                        if columns_to_drop:
                            joined_df = joined_df.drop(columns=columns_to_drop)
                
                # Format date columns
                for col in joined_df.columns:
                    if col.endswith('_at') and joined_df[col].dtype == 'object':
                        joined_df[col] = pd.to_datetime(joined_df[col])
                
                # Write the joined data to Excel
                if not joined_df.empty:
                    joined_df.to_excel(writer, sheet_name='Videos_With_Transcripts', index=False)
                    format_sheet(writer.sheets['Videos_With_Transcripts'], joined_df, 'Videos_With_Transcripts', is_main_view=True)
                    logger.debug(f"Exported joined view with {len(joined_df)} rows and columns: {', '.join(joined_df.columns)}")
                    
                # Create a keywords usage report
                if 'keywords' in inspector.get_table_names() and 'transcription_keywords' in inspector.get_table_names():
                    try:
                        # First try with GROUP_CONCAT
                        keywords_usage_query = """
                        SELECT k.name AS keyword, COUNT(tk.transcription_id) AS usage_count, 
                               GROUP_CONCAT(v.filename, ', ') AS videos
                        FROM keywords k
                        LEFT JOIN transcription_keywords tk ON k.id = tk.keyword_id
                        LEFT JOIN transcriptions t ON tk.transcription_id = t.id
                        LEFT JOIN videos v ON t.video_id = v.id
                        GROUP BY k.id
                        ORDER BY usage_count DESC, k.name
                        """
                        keywords_usage_df = pd.read_sql_query(keywords_usage_query, engine)
                    except Exception as e:
                        logger.debug(f"GROUP_CONCAT not supported in keywords usage, using alternative: {str(e)}")
                        # Get basic keyword usage count
                        base_query = """
                        SELECT k.name AS keyword, COUNT(tk.transcription_id) AS usage_count
                        FROM keywords k
                        LEFT JOIN transcription_keywords tk ON k.id = tk.keyword_id
                        GROUP BY k.id
                        ORDER BY usage_count DESC, k.name
                        """
                        keywords_usage_df = pd.read_sql_query(base_query, engine)
                        
                        # Get video filenames separately
                        if not keywords_usage_df.empty:
                            video_query = """
                            SELECT k.name AS keyword, v.filename
                            FROM keywords k
                            JOIN transcription_keywords tk ON k.id = tk.keyword_id
                            JOIN transcriptions t ON tk.transcription_id = t.id
                            JOIN videos v ON t.video_id = v.id
                            """
                            video_df = pd.read_sql_query(video_query, engine)
                            
                            # Aggregate videos by keyword
                            if not video_df.empty:
                                videos_agg = video_df.groupby('keyword')['filename'].apply(
                                    lambda x: ', '.join(x)
                                ).reset_index(name='videos')
                                
                                # Merge back to the keywords usage dataframe
                                keywords_usage_df = pd.merge(
                                    keywords_usage_df,
                                    videos_agg,
                                    on='keyword',
                                    how='left'
                                )
                    
                    if not keywords_usage_df.empty:
                        keywords_usage_df.to_excel(writer, sheet_name='Keywords_Usage', index=False)
                        format_sheet(writer.sheets['Keywords_Usage'], keywords_usage_df, 'Keywords_Usage', is_keywords=True)
                        logger.debug(f"Exported keywords usage report with {len(keywords_usage_df)} rows")
                    
            except Exception as e:
                logger.warning(f"Could not create joined view: {str(e)}")
                logger.debug(f"Error details: {str(e)}")
                # Try to create a basic joined view without the keywords if that was the issue
                try:
                    if not joined_df.empty:
                        # Still write the base joined data even if keywords failed
                        joined_df.to_excel(writer, sheet_name='Videos_With_Transcripts', index=False)
                        format_sheet(writer.sheets['Videos_With_Transcripts'], joined_df, 'Videos_With_Transcripts', is_main_view=True)
                        logger.info("Created basic joined view without keywords")
                except Exception as inner_e:
                    logger.warning(f"Could not create basic joined view either: {str(inner_e)}")
        
        logger.debug(f"Database exported to Excel: {excel_path}")
        
        return excel_path
    
    except Exception as e:
        logger.error(f"Error exporting database to Excel: {str(e)}")
        logger.debug(f"Error details: {str(e)}")
        return None 

def format_sheet(worksheet, df, table_name, is_main_view=False, is_keywords=False):
    """Format an Excel worksheet with styling and auto-width columns"""
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Define columns that need word wrap due to potentially long content
    long_content_columns = ['summary', 'transcript_text', 'keywords', 'videos']
    
    # Set column widths based on content
    for idx, col in enumerate(df.columns):
        column_letter = get_column_letter(idx + 1)
        
        # Determine if this is a long content column
        is_long_content = any(long_name in col.lower() for long_name in long_content_columns)
        
        if is_long_content:
            # For long content columns, set a fixed reasonable width
            worksheet.column_dimensions[column_letter].width = 50
        else:
            # For other columns, calculate width based on content
            max_length = max(
                df[col].astype(str).map(len).max(),  # max length of column content
                len(str(col))  # length of column header
            )
            adjusted_width = min(max_length + 2, 40)  # Limit max width to 40
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Style headers
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    
    # Style data cells and set word wrap for long content cells
    max_row_height = 60  # Maximum row height in points
    
    # Find the status column index for color coding
    status_idx = None
    for idx, col in enumerate(df.columns):
        if col.lower() == 'status':
            status_idx = idx
            break
    
    # Define status colors
    status_colors = {
        "Missing": "FF9999",  # Light red
        "Error Transcribing": "FFCC99",  # Light orange
        "New": "FFFFCC",  # Light yellow
        "Transcribed": "CCFFCC",  # Light green
    }
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row)):
        # Set a default row height (will be overridden if needed)
        worksheet.row_dimensions[row_idx + 2].height = 15  # Standard height
        
        # Track if this row needs height adjustment for long content
        row_has_long_content = False
        
        for cell_idx, cell in enumerate(row):
            # Get the column name for this cell
            col_name = df.columns[cell_idx]
            
            # Check if this is a long content column
            is_long_content = any(long_name in col_name.lower() for long_name in long_content_columns)
            
            if is_long_content and cell.value:
                # Apply word wrap and vertical alignment for long content cells
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                row_has_long_content = True
            else:
                # For other cells, center short content
                cell.alignment = Alignment(vertical="center")
            
            # Apply status color coding if this is the status column
            if status_idx is not None and cell_idx == status_idx and cell.value in status_colors:
                cell.fill = PatternFill(
                    start_color=status_colors[cell.value],
                    end_color=status_colors[cell.value],
                    fill_type="solid"
                )
                
                # Make missing status bold and red text
                if cell.value == "Missing":
                    cell.font = Font(bold=True, color="990000")
                # Make error status bold
                elif cell.value == "Error Transcribing":
                    cell.font = Font(bold=True)
                
        # If the row has long content, set a taller but limited row height
        if row_has_long_content:
            worksheet.row_dimensions[row_idx + 2].height = max_row_height
            
        # If this is a missing or error file, highlight the entire row with a light tint
        if status_idx is not None and row[status_idx].value in ["Missing", "Error Transcribing"]:
            highlight_color = "FFEEEE" if row[status_idx].value == "Missing" else "FFF6EE"
            for cell in row:
                # Don't override the status cell which already has its own color
                if cell_idx != status_idx and not cell.fill.start_color.index.startswith('FF'):
                    cell.fill = PatternFill(
                        start_color=highlight_color,
                        end_color=highlight_color,
                        fill_type="solid"
                    )
    
    # Format as a table with filtering
    data_range = worksheet.dimensions
    table = Table(displayName=f"Table_{table_name.replace(' ', '_')}", ref=data_range)
    
    # Choose an appropriate table style
    if is_main_view:
        style = "TableStyleMedium2"  # Blue style for main view
    elif is_keywords:
        style = "TableStyleMedium7"  # Orange style for keywords
    else:
        style = "TableStyleMedium9"  # Default style
        
    table_style = TableStyleInfo(
        name=style, 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=False
    )
    table.tableStyleInfo = table_style
    
    # Add the table to the worksheet
    worksheet.add_table(table)
    
    # Freeze the header row
    worksheet.freeze_panes = 'A2'

def create_metadata_sheet(writer):
    """Create a metadata/help sheet with field descriptions"""
    # Create a new sheet
    workbook = writer.book
    metadata_sheet = workbook.create_sheet("Info", 0)  # Add at the beginning
    
    # Set as active sheet
    workbook.active = 0
    
    # Define styles
    title_font = Font(size=14, bold=True, color="1F4E78")
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    # Add title
    metadata_sheet['A1'] = "Video Library Database Information"
    metadata_sheet['A1'].font = title_font
    metadata_sheet.merge_cells('A1:C1')
    metadata_sheet['A1'].alignment = Alignment(horizontal="center")
    
    # Add creation date
    metadata_sheet['A3'] = "Export Date:"
    metadata_sheet['B3'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Sheet description
    metadata_sheet['A5'] = "This workbook contains the following sheets:"
    metadata_sheet['A5'].font = header_font
    
    sheet_descriptions = [
        ("Info", "This information sheet with field descriptions and help"),
        ("Videos_With_Transcripts", "Main view with video metadata, transcriptions, and keywords"),
        ("keywords", "List of all keywords used in the system"),
        ("videos", "Raw video metadata"),
        ("transcriptions", "Raw transcription data"),
        ("Keywords_Usage", "Keyword usage statistics")
    ]
    
    for i, (sheet_name, desc) in enumerate(sheet_descriptions):
        row = 6 + i
        metadata_sheet[f'A{row}'] = sheet_name
        metadata_sheet[f'B{row}'] = desc
    
    # Add field descriptions
    metadata_sheet['A12'] = "Field Descriptions:"
    metadata_sheet['A12'].font = header_font
    
    # Column headers
    metadata_sheet['A13'] = "Field Name"
    metadata_sheet['B13'] = "Description"
    metadata_sheet['C13'] = "Data Type"
    
    # Style the headers
    for cell in ['A13', 'B13', 'C13']:
        metadata_sheet[cell].font = header_font
        metadata_sheet[cell].fill = header_fill
    
    # List of field descriptions
    field_descriptions = [
        ("id", "Unique identifier for each record", "Integer"),
        ("filename", "Name of the video file", "Text"),
        ("filepath", "Full path to the video file", "Text"),
        ("filesize", "Size of the file in bytes", "Integer"),
        ("duration", "Duration of the video in seconds", "Decimal"),
        ("resolution", "Video resolution (width x height)", "Text"),
        ("width", "Video width in pixels", "Integer"),
        ("height", "Video height in pixels", "Integer"),
        ("encoding", "Video codec/encoding format", "Text"),
        ("bitrate", "Video bitrate in bits per second", "Integer"),
        ("fps", "Frames per second", "Decimal"),
        ("status", "Current status of the video (New, Transcribed, Missing, Error Transcribing)", "Text"),
        ("is_transcribed", "Whether the video has been transcribed", "Boolean"),
        ("transcribed_at", "Date and time when the video was transcribed", "Date/Time"),
        ("transcript_text", "Full text transcript of the video", "Text"),
        ("transcript_file", "Path to the transcript file", "Text"),
        ("suggested_title", "AI-generated title for the video", "Text"),
        ("summary", "AI-generated summary of the video content", "Text"),
        ("keywords", "Keywords related to the video content", "Text (comma-separated)"),
        ("created_at", "Date and time when the record was created", "Date/Time"),
        ("updated_at", "Date and time when the record was last updated", "Date/Time"),
        ("usage_count", "Number of times a keyword is used across all videos", "Integer")
    ]
    
    # Add field descriptions
    for i, (field, desc, data_type) in enumerate(field_descriptions):
        row = 14 + i
        metadata_sheet[f'A{row}'] = field
        metadata_sheet[f'B{row}'] = desc
        metadata_sheet[f'C{row}'] = data_type
    
    # Set column widths
    metadata_sheet.column_dimensions['A'].width = 20
    metadata_sheet.column_dimensions['B'].width = 60
    metadata_sheet.column_dimensions['C'].width = 15
    
    # Add usage instructions
    instruction_row = 14 + len(field_descriptions) + 2
    metadata_sheet[f'A{instruction_row}'] = "Usage Tips:"
    metadata_sheet[f'A{instruction_row}'].font = header_font
    
    tips = [
        "Use the 'Videos_With_Transcripts' sheet for most operations, as it combines all relevant data.",
        "Click on the filter buttons in column headers to sort or filter data.",
        "The 'Keywords_Usage' sheet shows how frequently each keyword is used.",
        "Keywords help categorize videos and can be used to find related content.",
        "The 'status' column shows each video's current state: New, Transcribed, Missing, or Error Transcribing.",
        "Missing videos (red highlight) means the file was once in the database but can no longer be found.",
        "Sort by status to quickly identify videos that need attention or are ready for use."
    ]
    
    for i, tip in enumerate(tips):
        row = instruction_row + 1 + i
        metadata_sheet[f'A{row}'] = f"• {tip}"
        metadata_sheet.merge_cells(f'A{row}:C{row}')
        metadata_sheet[f'A{row}'].alignment = Alignment(wrap_text=True)
    
    # Set row height for wrapped text
    for row in range(instruction_row + 1, instruction_row + 1 + len(tips)):
        metadata_sheet.row_dimensions[row].height = 30 